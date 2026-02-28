"""Excel export with 4 formatted sheets using openpyxl."""

from collections import Counter
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
 
# ═══════════════════════════════════════════
# COLOR DEFINITIONS
# ═══════════════════════════════════════════
RED_FILL = PatternFill(start_color="FFD6D6", end_color="FFD6D6", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF9D6", end_color="FFF9D6", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
STAT_FONT = Font(name="Calibri", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# ═══════════════════════════════════════════
# COLUMN DEFINITIONS
# ═══════════════════════════════════════════
ALL_LEADS_COLUMNS = [
    "Business Name", "Category / Type", "Owner Name", "Phone Number",
    "Email Address", "Full Address", "Locality / Area", "Google Maps Link",
    "Has Website?", "Website URL", "Google Rating", "Number of Reviews",
    "Instagram Link", "Facebook Link", "Claimed on Google?", "Date Added",
    "Priority Score", "Lead Status", "Notes",
    "WA Message (Short)", "WA Message (Full)", "WA Ready Link",
]

HOT_LEADS_COLUMNS = [
    "Business Name", "Category", "Phone", "Address",
    "WA Short Message", "WA Ready Link",
]

OUTREACH_COLUMNS = [
    "Business Name", "Category", "Phone Number", "Locality",
    "Short Message", "Full Message", "WA Ready Link",
]


def export_to_excel(businesses: list[dict], location: str, filename: str) -> str:
    """Create the formatted Excel file with 4 sheets."""
    wb = Workbook()

    _create_all_leads_sheet(wb, businesses)
    _create_hot_leads_sheet(wb, businesses)
    _create_outreach_sheet(wb, businesses)
    _create_summary_sheet(wb, businesses, location)

    wb.save(filename)
    print(f"\n📁 Excel file saved: {filename}")
    return filename


def _style_header_row(ws, columns):
    """Apply header styling to the first row."""
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"


def _auto_width(ws, columns, max_width=40):
    """Set reasonable column widths."""
    for col_idx, col_name in enumerate(columns, 1):
        width = min(len(col_name) + 6, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _row_fill(lead_status: str, phone: str):
    """Return the appropriate fill for a row."""
    if not phone:
        return ORANGE_FILL
    if lead_status == "HOT":
        return RED_FILL
    if lead_status == "WARM":
        return YELLOW_FILL
    return None


def _biz_to_all_leads_row(biz: dict) -> list:
    """Convert a business dict to a row for the All Leads sheet."""
    return [
        biz.get("name", ""),
        biz.get("category_label", ""),
        biz.get("owner_name", ""),
        biz.get("phone_normalized", "") or biz.get("phone", ""),
        biz.get("email", ""),
        biz.get("address", ""),
        biz.get("locality", ""),
        biz.get("google_maps_link", ""),
        "Yes" if biz.get("has_website") else "No",
        biz.get("website", ""),
        biz.get("google_rating", ""),
        biz.get("review_count", 0),
        biz.get("instagram", ""),
        biz.get("facebook", ""),
        biz.get("claimed", "Unknown"),
        date.today().isoformat(),
        biz.get("priority_score", 0),
        biz.get("lead_status", ""),
        biz.get("notes", ""),
        biz.get("wa_short", ""),
        biz.get("wa_full", ""),
        biz.get("wa_link", ""),
    ]


def _create_all_leads_sheet(wb: Workbook, businesses: list[dict]):
    """Sheet 1: All Leads — full 22-column data."""
    ws = wb.active
    ws.title = "All Leads"

    _style_header_row(ws, ALL_LEADS_COLUMNS)
    _auto_width(ws, ALL_LEADS_COLUMNS)

    for row_idx, biz in enumerate(businesses, 2):
        row_data = _biz_to_all_leads_row(biz)
        fill = _row_fill(biz.get("lead_status", ""), biz.get("phone_normalized", ""))

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN
            if fill:
                cell.fill = fill


def _create_hot_leads_sheet(wb: Workbook, businesses: list[dict]):
    """Sheet 2: Hot Leads — score >= 8 only."""
    ws = wb.create_sheet("Hot Leads")

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "🔴 HOT LEADS — Contact These First"
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER_ALIGN

    for col_idx, col_name in enumerate(HOT_LEADS_COLUMNS, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    ws.freeze_panes = "A3"
    _auto_width(ws, HOT_LEADS_COLUMNS)

    hot = [b for b in businesses if b.get("lead_status") == "HOT"]
    for row_idx, biz in enumerate(hot, 3):
        row_data = [
            biz.get("name", ""),
            biz.get("category_label", ""),
            biz.get("phone_normalized", "") or biz.get("phone", ""),
            biz.get("address", ""),
            biz.get("wa_short", ""),
            biz.get("wa_link", ""),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.fill = RED_FILL
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN


def _create_outreach_sheet(wb: Workbook, businesses: list[dict]):
    """Sheet 3: Outreach Messages — all businesses with messages."""
    ws = wb.create_sheet("Outreach Messages")

    total = len(businesses)
    wa_ready = sum(1 for b in businesses if b.get("wa_link", "").startswith("https://"))
    missing_phone = sum(1 for b in businesses if b.get("wa_link") == "PHONE MISSING")
    hot_count = sum(1 for b in businesses if b.get("lead_status") == "HOT")

    summary_lines = [
        f"Total Messages Generated     : {total}",
        f"WhatsApp Links Ready         : {wa_ready}",
        f"Missing Phone Numbers        : {missing_phone}",
        f"Hot Leads to Contact Today   : {hot_count}",
    ]

    for i, line in enumerate(summary_lines):
        cell = ws.cell(row=i + 1, column=1, value=line)
        cell.font = STAT_FONT

    header_row = len(summary_lines) + 2
    for col_idx, col_name in enumerate(OUTREACH_COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    ws.freeze_panes = f"A{header_row + 1}"
    _auto_width(ws, OUTREACH_COLUMNS)

    ws.column_dimensions["E"].width = 35
    ws.column_dimensions["F"].width = 50

    for row_idx, biz in enumerate(businesses, header_row + 1):
        phone = biz.get("phone_normalized", "")
        row_data = [
            biz.get("name", ""),
            biz.get("category_label", ""),
            phone or biz.get("phone", ""),
            biz.get("locality", ""),
            biz.get("wa_short", ""),
            biz.get("wa_full", ""),
            biz.get("wa_link", ""),
        ]
        fill = ORANGE_FILL if not phone else None
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN
            if fill:
                cell.fill = fill


def _create_summary_sheet(wb: Workbook, businesses: list[dict], location: str):
    """Sheet 4: Summary & Stats."""
    ws = wb.create_sheet("Summary & Stats")

    total = len(businesses)
    no_website = sum(1 for b in businesses if not b.get("has_website"))
    hot = sum(1 for b in businesses if b.get("lead_status") == "HOT")
    warm = sum(1 for b in businesses if b.get("lead_status") == "WARM")
    cold = sum(1 for b in businesses if b.get("lead_status") == "COLD")
    wa_ready = sum(1 for b in businesses if b.get("wa_link", "").startswith("https://"))
    missing_phone = sum(1 for b in businesses if b.get("wa_link") == "PHONE MISSING")

    cat_counts = Counter(b.get("category_label", "") for b in businesses)
    top_3_cats = [cat for cat, _ in cat_counts.most_common(3)]

    top_5 = [b for b in businesses if b.get("lead_status") == "HOT"][:5]

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 50

    ws.merge_cells("A1:B1")
    title = ws["A1"]
    title.value = "📊 Lead Finder — Summary & Stats"
    title.font = TITLE_FONT

    stats = [
        ("Total Businesses Found", total),
        ("Total with No Website", no_website),
        ("Total HOT Leads 🔴", hot),
        ("Total WARM Leads 🟡", warm),
        ("Total COLD Leads 🟢", cold),
        ("Total WA Links Ready", wa_ready),
        ("Total Missing Phone Numbers", missing_phone),
        ("Top 3 Categories Found", ", ".join(top_3_cats)),
        ("Date of Last Run", date.today().isoformat()),
        ("Location Searched", location),
    ]

    for i, (label, value) in enumerate(stats, 3):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font = STAT_FONT
        label_cell.border = THIN_BORDER

        val_cell = ws.cell(row=i, column=2, value=str(value))
        val_cell.font = BODY_FONT
        val_cell.border = THIN_BORDER

    top5_start = len(stats) + 5
    ws.cell(row=top5_start, column=1, value="🏆 Top 5 Businesses to Call TODAY").font = TITLE_FONT

    for col_idx, header in enumerate(["#", "Business Name", "Phone", "Score"], 1):
        cell = ws.cell(row=top5_start + 1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    for i, biz in enumerate(top_5, 1):
        row = top5_start + 1 + i
        ws.cell(row=row, column=1, value=i).border = THIN_BORDER
        ws.cell(row=row, column=2, value=biz.get("name", "")).border = THIN_BORDER
        ws.cell(row=row, column=3, value=biz.get("phone_normalized", "")).border = THIN_BORDER
        ws.cell(row=row, column=4, value=biz.get("priority_score", 0)).border = THIN_BORDER

        for col in range(1, 5):
            ws.cell(row=row, column=col).fill = RED_FILL
